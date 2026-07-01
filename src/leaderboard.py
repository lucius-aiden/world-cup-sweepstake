from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import LeaderboardRow, MovementRecord

BLACK_FILL = PatternFill(fill_type="solid", fgColor="0A0A0A")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="050505")
ALIVE_FILL = PatternFill(fill_type="solid", fgColor="0A0A0A")
ELIMINATED_FILL = PatternFill(fill_type="solid", fgColor="161616")
GOLD = "E0B400"
WHITE = "F8F8F8"
GRAY = "B7B7B7"
BORDER = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="thin", color=GOLD),
)


def build_leaderboard(
    rows: list[dict],
    previous_ranks: dict[str, int],
    previous_points: dict[str, int] | None = None,
) -> tuple[list[LeaderboardRow], list[MovementRecord]]:
    previous_points = previous_points or {}
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[row["player"]].append(dict(row))

    leaderboard: list[LeaderboardRow] = []
    for player, teams in grouped.items():
        if len(teams) != 2:
            raise ValueError(f"Expected exactly 2 teams for {player}, found {len(teams)}")
        teams = sorted(teams, key=lambda item: int(item.get("team_slot", 0)))
        
        first, second = teams[0], teams[1]
        first_points = int(first["points"])
        second_points = int(second["points"])
        first_score = int(first.get("contributing_points", first_points)) + int(first.get("advancement_bonus", 0))
        second_score = int(second.get("contributing_points", second_points)) + int(second.get("advancement_bonus", 0))
        leaderboard.append(
            LeaderboardRow(
                player=player,
                team_1=first["team_name"],
                team_1_points=first_points,
                team_1_status=_status_label(_alive_flag(first.get("alive", 0))),
                team_2=second["team_name"],
                team_2_points=second_points,
                team_2_status=_status_label(_alive_flag(second.get("alive", 0))),
                total_points=first_score + second_score,
                teams_alive=int(_alive_flag(first.get("alive", 0))) + int(_alive_flag(second.get("alive", 0))),
                rank=0,
            )
        )

    leaderboard.sort(
        key=lambda row: (
            -row.total_points,
            -row.teams_alive,
            row.player.lower(),
        )
    )
    ranked_rows: list[LeaderboardRow] = []
    movements: list[MovementRecord] = []
    for index, row in enumerate(leaderboard, start=1):
        ranked = LeaderboardRow(**{**row.__dict__, "rank": index})
        ranked_rows.append(ranked)
        previous_rank = previous_ranks.get(ranked.player)
        movements.append(
            MovementRecord(
                player=ranked.player,
                previous_rank=previous_rank,
                new_rank=index,
                previous_points=previous_points.get(ranked.player),
                current_points=ranked.total_points,
            )
        )
    return ranked_rows, movements


def write_workbook(leaderboard: list[LeaderboardRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leaderboard"

    headers = [
        "Rank",
        "Player",
        "Team 1",
        "Team 1 Points",
        "Team 1 Status",
        "Team 2",
        "Team 2 Points",
        "Team 2 Status",
        "Total Points",
        "Teams Alive",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color=WHITE, size=14)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in leaderboard:
        sheet.append(
            [
                row.rank,
                row.player,
                row.team_1,
                row.team_1_points,
                row.team_1_status,
                row.team_2,
                row.team_2_points,
                row.team_2_status,
                row.total_points,
                row.teams_alive,
            ]
        )

    for data_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in data_row:
            cell.fill = BLACK_FILL
            cell.font = Font(color=WHITE, size=12)
            cell.border = BORDER
            if cell.column in {1, 4, 7, 9, 10}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_status_fill(data_row[4], data_row[4].value)
        _apply_status_fill(data_row[7], data_row[7].value)
        data_row[4].font = Font(color=WHITE, bold=True)
        data_row[7].font = Font(color=WHITE, bold=True)

    for column_letter in ("C", "F"):
        for cell in sheet[column_letter][1:]:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    _autosize(sheet)
    workbook.save(output_path)


def _status_label(alive: bool) -> str:
    return "Still in" if alive else "Knocked out"


def _alive_flag(value: object) -> bool:
    if isinstance(value, str):
        cleaned = value.strip().lower()
        if cleaned in {"", "0", "false", "no", "n"}:
            return False
        if cleaned in {"1", "true", "yes", "y"}:
            return True
    return bool(value)


def _apply_status_fill(cell, value: str) -> None:
    cell.fill = ALIVE_FILL if value == "Still in" else ELIMINATED_FILL
    cell.font = Font(color=WHITE if value == "Still in" else GRAY, bold=True)


def _autosize(sheet) -> None:
    for column in sheet.columns:
        width = max(len("" if cell.value is None else str(cell.value)) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = width
